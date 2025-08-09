from flask import Blueprint, Response, request
from flask_jwt_extended import jwt_required
from datetime import date
from openpyxl import Workbook
from backend.services.kpi import get_kpi_snapshot
from backend.services.meta_campaign import get_meta_campaign_rows
from backend.services.bm_profit import get_profit_summary
from backend.services.portfolio import get_bm_health_rows
from backend.services.campaign_command import get_campaign_command_totals
from backend.utils.db import get_db_session
from backend.utils.excel import build_excel_response, _build_campaign_command_workbook
from backend.utils.pagination import paginate

export_bp = Blueprint('export', __name__, url_prefix='/api/v1/export')

@export_bp.route('/kpi_snapshot.xlsx', methods=['GET'])
@jwt_required()
async def export_kpi_snapshot():
    start_date = date.fromisoformat(request.args.get('start_date'))
    end_date = date.fromisoformat(request.args.get('end_date'))
    master_store_ids = request.args.getlist('master_store_ids')
    bm_ids = request.args.getlist('bm_ids')
    mode = request.args.get('mode', 'native')
    async with get_db_session() as session:
        rows = await get_kpi_snapshot(session, start_date, end_date, master_store_ids, bm_ids, mode)
        wb = Workbook()
        ws = wb.active
        ws.append(['Date', 'Revenue', 'Ad Spend', 'ROAS', 'CPA', 'Currency'])
        for row in rows:
            ws.append([row.date, row.revenue, row.ad_spend, row.roas, row.cpa, row.currency_code])
        return build_excel_response(wb, 'kpi_snapshot.xlsx')

@export_bp.route('/campaign_command.xlsx', methods=['GET'])
@jwt_required()
async def export_campaign_command():
    start_date = date.fromisoformat(request.args.get('start_date'))
    end_date = date.fromisoformat(request.args.get('end_date'))
    filters = {k: request.args.getlist(k) for k in ('master_store_ids', 'bm_ids', 'status')}
    mode = request.args.get('mode', 'native')
    async with get_db_session() as session:
        stmt = await get_campaign_command_data(session, start_date, end_date, filters)
        rows = (await session.execute(stmt)).all()
        totals = await get_campaign_command_totals(session, start_date, end_date, filters)
        wb = _build_campaign_command_workbook([r._asdict() for r in rows], totals._asdict(), request.args)
        return build_excel_response(wb, 'campaign_command.xlsx')

@export_bp.route('/meta_campaign_data.xlsx', methods=['GET'])
@jwt_required()
async def export_meta_campaign_data():
    start_date = date.fromisoformat(request.args.get('start_date'))
    end_date = date.fromisoformat(request.args.get('end_date'))
    filters = {k: request.args.getlist(k) for k in ('bm_ids', 'campaign_ids')}
    async with get_db_session() as session:
        rows = await get_meta_campaign_rows(session, start_date, end_date, filters)
        wb = Workbook()
        ws = wb.active
        ws.append(['Campaign ID', 'Date', 'Name', 'Status', 'Ad Budget', 'Reach', 'Landing Page Views'])
        for row in rows:
            ws.append([row.campaign_id, row.date, row.name, row.status, row.ad_budget, row.reach, row.landing_page_views])
        return build_excel_response(wb, 'meta_campaign_data.xlsx')

@export_bp.route('/bm_profitability.xlsx', methods=['GET'])
@jwt_required()
async def export_bm_profitability():
    start_date = date.fromisoformat(request.args.get('start_date'))
    end_date = date.fromisoformat(request.args.get('end_date'))
    bm_ids = request.args.getlist('bm_ids')
    mode = request.args.get('mode', 'native')
    async with get_db_session() as session:
        rows = await get_profit_summary(session, start_date, end_date, bm_ids)
        wb = Workbook()
        ws = wb.active
        ws.append(['BM ID', 'BM Name', 'Revenue', 'Ad Spend', 'Profit Margin %', 'Fixed Costs', 'Variable Costs %', 'Net Profit', 'Adjusted ROAS'])
        for row in rows:
            ws.append([row.bm_id, row.bm_name, row.revenue, row.ad_spend, row.profit_margin_pct, row.fixed_costs, row.variable_costs_pct, row.net_profit, row.adjusted_roas])
        return build_excel_response(wb, 'bm_profitability.xlsx')

@export_bp.route('/portfolio_snapshot.xlsx', methods=['GET'])
@jwt_required()
async def export_portfolio_snapshot():
    region_ids = request.args.getlist('region_ids')
    master_store_ids = request.args.getlist('master_store_ids')
    async with get_db_session() as session:
        rows = await get_bm_health_rows(session, region_ids, master_store_ids)
        wb = Workbook()
        ws = wb.active
        ws.append(['BM ID', 'BM Name', 'Master Store ID', 'Last Meta Fetch', 'Last Shopify Fetch', 'Token Status', 'Active', 'Age Meta (hours)', 'Age Shopify (hours)'])
        for row in rows:
            ws.append([row.bm_id, row.bm_name, row.master_store_id, row.last_successful_fetch_meta_at, row.last_successful_fetch_shopify_at, row.meta_token_status, row.is_active, row.age_meta_hours, row.age_shopify_hours])
        return build_excel_response(wb, 'portfolio_snapshot.xlsx')