import pytest
import openpyxl
import io
from datetime import date, timedelta
from decimal import Decimal
from flask.testing import FlaskClient
from sqlalchemy.ext.asyncio import AsyncSession
from uuid import uuid4

from backend.models import (
    Region, MasterStoreConfig, BusinessManagerConfig, ProductCategory,
    FXDailyRate, User
)


# Helper from unit tests
def get_auth_headers(client: FlaskClient, email: str, password: str = "password") -> dict:
    rv = client.post("/api/v1/auth/login", json={"email": email, "password": password})
    access_token = rv.json["access_token"]
    return {"Authorization": f"Bearer {access_token}"}


@pytest.fixture(scope="module")
def celery_eager_app(app):
    """Fixture to configure the app for eager Celery task execution."""
    app.config.update({"CELERY_TASK_ALWAYS_EAGER": True})
    yield app


@pytest.fixture(scope="function")
async def setup_full_system_config(db_session: AsyncSession):
    """Sets up a complete chain of config objects for an end-to-end test."""
    admin_user = User(email="integ_admin@test.com", password_hash=current_app.bcrypt.generate_password_hash("password"), is_admin=True, is_active=True)
    region = Region(name="EU", currency_code="EUR", is_default=True, is_active=True)
    category = ProductCategory(id=uuid4(), name="Test Category", slug="test-cat", is_default=True, is_active=True)

    db_session.add_all([admin_user, region, category])
    await db_session.commit()

    master_store = MasterStoreConfig(
        id=uuid4(),
        name="EU Master Store",
        shopify_domain="eu-master.myshopify.com",
        store_type=StoreType.MASTER,
        region_id=region.id,
        is_active=True
    )
    db_session.add(master_store)
    await db_session.commit()

    bm_config = BusinessManagerConfig(
        master_store_id=master_store.id,
        name="EU Business Manager",
        child_shopify_domain="eu-child.myshopify.com",
        meta_bm_id="bm_12345",
        current_product_category_id=category.id,
        is_active=True,
        meta_token_status=MetaTokenStatus.ACTIVE
    )
    db_session.add(bm_config)
    await db_session.commit()

    test_date = date.today() - timedelta(days=1)

    auto_rate = FXDailyRate(
        date=test_date, from_currency="EUR", to_currency="USD",
        rate=Decimal("1.10"), source="exchangerate.host"
    )
    manual_rate = FXDailyRate(
        date=test_date, from_currency="EUR", to_currency="USD",
        rate=Decimal("1.50"), source="manual"
    )
    db_session.add_all([auto_rate, manual_rate])
    await db_session.commit()

    return {
        "admin_user": admin_user,
        "bm_config": bm_config,
        "master_store": master_store,
        "test_date": test_date
    }


@pytest.mark.asyncio
async def test_ingest_to_usd_export_critical_path(
    celery_eager_app, client: FlaskClient, db_session: AsyncSession, setup_full_system_config
):
    config = await setup_full_system_config
    bm_id = config["bm_config"].id
    test_date_str = config["test_date"].isoformat()
    ingestion_token = celery_eager_app.config['INGESTION_TOKEN']

    ingest_payload = {
        "bm_id": bm_id,
        "date": test_date_str,
        "data_type": "meta",
        "payload": [
            {"campaign_id": "c1", "spend": 100.00, "clicks": 50, "impressions": 1000}
        ]
    }
    ingest_headers = {"X-Ingestion-Token": ingestion_token}
    rv = client.post("/api/v1/ingest", json=ingest_payload, headers=ingest_headers)
    assert rv.status_code == 202

    ingest_payload_shopify = {
        "bm_id": bm_id,
        "date": test_date_str,
        "data_type": "shopify_child",
        "payload": {"orders_count": 10, "gross_sales": 1000.00}
    }
    rv = client.post("/api/v1/ingest", json=ingest_payload_shopify, headers=ingest_headers)
    assert rv.status_code == 202

    from backend.models.aggregated import KPIDailySnapshot

    snapshot = (await db_session.execute(select(KPIDailySnapshot).filter_by(
        bm_id=bm_id, date=config["test_date"]
    ))).scalars().first()

    assert snapshot is not None
    assert snapshot.currency_code == "USD"  # Adjusted
    assert snapshot.ad_spend == Decimal("100.00")
    assert snapshot.revenue == Decimal("1000.00")

    admin_headers = get_auth_headers(client, config["admin_user"].email)
    export_url = f"/api/v1/export/kpi_snapshot.xlsx?bm_ids={bm_id}&mode=usd"
    rv = client.get(export_url, headers=admin_headers)

    assert rv.status_code == 200
    assert rv.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = openpyxl.load_workbook(io.BytesIO(rv.data))
    sheet = workbook.active

    header_row = [cell.value for cell in sheet[1]]
    spend_col_idx = header_row.index("Ad Spend") + 1
    revenue_col_idx = header_row.index("Revenue") + 1

    spend_in_usd = sheet.cell(row=2, column=spend_col_idx).value
    revenue_in_usd = sheet.cell(row=2, column=revenue_col_idx).value

    assert spend_in_usd == pytest.approx(150.00)
    assert revenue_in_usd == pytest.approx(1500.00)